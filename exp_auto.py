import pandas as pd
import numpy as np
import os
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

EXPENSES_PATH = r"C:\Users\yaron\שולחן העבודה\expenses.xlsx"
FOLDER_PATH = r"C:\Users\yaron\שולחן העבודה\תיקיה אישית\הוצאות"
SUBSCRIPTIONS = ['SPOTIFYIL', 'APPLE', 'פיס מנויים']
SAVING_DEPOSITS_DIC = {"אנליסט": 200}

def date_change(row):
    # Changes the date for transactions with multiple payments
    if not pd.isna(row['הערות']):
        if re.search(r"תשלום \d* מתוך \d*", row['הערות']):
            #If the row in the pattern, extract the payment number (don't change if it's the first payment):
            add_to_month = int(re.search(r'\d+', row['הערות']).group()) - 1
            if add_to_month != 0:
                date_parts = row['תאריך'].split('-')
                new_month = int(date_parts[1]) + add_to_month
                if new_month <= 9: #1-9
                    row['תאריך'] = f"10-0{str(new_month)}-{date_parts[2]}"
                elif new_month <= 12: #10-12
                    row['תאריך'] = f"10-{str(new_month)}-{date_parts[2]}"
                else: #13+
                    if new_month % 12 <= 9:
                        row['תאריך'] = f"10-0{str(new_month % 12)}-{str(int(date_parts[2]) + new_month // 12)}"
                    else:
                        row['תאריך'] = f"10-{str(new_month % 12)}-{str(int(date_parts[2]) + new_month // 12)}"
    return row

def add_deposits(df):
    # Function that adds the savings deposits for new months
    first_date = df.iloc[0]["תאריך"]
    month = first_date[3:5]
    year = first_date[6:]
    for business,deposit in SAVING_DEPOSITS_DIC.items():
        new_row = [f"10-{month}-{year}","חיסכון",business,"חיסכון",'עו"ש',deposit, None, deposit]
        df.loc[len(df)] = new_row
    return df


def edit_month_file(month_path):
    # Re-orginaize files of expenses by the relevant columns and the re-named columns.
    israel_expenses = pd.read_excel(month_path, sheet_name=0, header=3, skipfooter=3, usecols=range(11))
    abroad_expenses = pd.read_excel(month_path, sheet_name='עסקאות חו"ל ומט"ח', header=3, skipfooter=3, usecols=range(11))

    COLS_ORDER = ["תאריך עסקה", "קטגוריה", "שם בית העסק", "סוג עסקה", "4 ספרות אחרונות של כרטיס האשראי",
                  "סכום עסקה מקורי", "הערות", "סכום חיוב"]
    COLS_RENAMED = ["תאריך", "קטגוריה", "שם בית העסק", "סוג עסקה", "כרטיס", "סכום כולל", "הערות", "חיוב"]

    all_expenses = pd.concat([israel_expenses, abroad_expenses], ignore_index=True)[COLS_ORDER]
    all_expenses.columns = COLS_RENAMED

    #Update the dates for transactions with multiple payments
    all_expenses = all_expenses.apply(date_change, axis=1)

    #Add the regular deposits
    all_expenses = add_deposits(all_expenses)

    return my_style(all_expenses)

def get_new_rows(main_df, new_df):
    #Gets 2 df, and returns the expenses that appears only at the second df.
    merged = pd.merge(new_df, main_df, how='left', indicator=True)
    new_rows = merged[merged['_merge'] == 'left_only']
    return new_rows

def everything_added(main_df, new_df):
    # Checks if all the expenses are already at the main expenses file.
    new_rows = get_new_rows(main_df, new_df)
    if len(new_rows) == 0:
        print(f"There are 0 new expenses.")
        return True
    else:
        print(f"There are {len(new_rows)} new expenses: ")
        print(new_rows[["תאריך", "שם בית העסק", "הערות", "חיוב"]].set_index("תאריך"))
        answer = input("To add press 1: ")
        if answer == "1": return False
        return True


def add_file(expenses_path, new):
    # Adds the new expenses to the expenses file, and ensures the data types
    # of the relevant columns
    ILS_FORMAT = '_ [$₪-he-IL] * #,##0.00_ ;_ [$₪-he-IL] * -#,##0.00_ ;_ [$₪-he-IL] * "-"??_ ;_ @_ '
    expenses_table = pd.read_excel(expenses_path, sheet_name=0)
    wb = openpyxl.load_workbook(expenses_path)
    ws = wb['exp']

    # Table reshape
    table_range = f'A1:H{len(expenses_table) + len(new) + 1}'
    table = ws._tables['exp']
    table.ref = table_range

    # Adding the new expenses
    for row in dataframe_to_rows(new, index=False, header=False):
        ws.append(row)
    #Ensure the Excel columns types
    for cell in ws['A']:
        cell.number_format = numbers.FORMAT_DATE_DDMMYY

    for cell in ws['F']:
        cell.number_format = ILS_FORMAT

    for cell in ws['H']:
        cell.number_format = ILS_FORMAT

    wb.save(expenses_path)

def english_expenses_name_change(df):
    # Changes the businesses name for the intenational expenses
    shop_name = "שם בית העסק"

    # The pattern to get the first word:
    pattern = re.compile(r"^[a-zA-Z]+")

    # Gets the relevant rows and make the change:
    new_english_rows = df.loc[df[shop_name].str.contains(pattern), shop_name].str.extract \
        (pat=r"(^[a-zA-Z]+)", expand=False)

    # Now change the df:
    df.loc[df[shop_name].str.contains(pattern), shop_name] = new_english_rows
    return df

def my_style(df):
    # Defines private style for every new df of expenses, by various rules.
    regular = "רגילה"
    subscribe = "מינוי"
    food = 'אוכל ושתייה'
    cloth = 'ביגוד'
    staff = 'שונות'

    df['תאריך'] = pd.to_datetime(df['תאריך'], dayfirst=True)
    df = english_expenses_name_change(df)

    df.loc[df["הערות"].str.contains('חיוב עסקת חו"ל', na=True), "הערות"] = None
    df['סוג עסקה'] = df['סוג עסקה'].replace({'חיוב חודשי': regular})
    df['סוג עסקה'] = np.where(df["שם בית העסק"].isin(SUBSCRIPTIONS), subscribe, df['סוג עסקה'])

    df['קטגוריה'] = df['קטגוריה'].replace(
        {'מזון וצריכה': food, 'הלבשה והנעלה': cloth, 'כלבו': staff, "ספרים והוצ' משרד": staff})

    return df


def main():
    #Let's start!
    main_df = pd.read_excel(EXPENSES_PATH, sheet_name=0)
    month = input("Insert a month of expenses:  ")
    file_name = f"{month}.xlsx"
    file_path = os.path.join(FOLDER_PATH, file_name)
    try:
        full_new_df = edit_month_file(file_path)
        if not everything_added(main_df, full_new_df):
            new_rows = get_new_rows(main_df, full_new_df)
            add_file(EXPENSES_PATH, full_new_df)
            print("Done")

    except FileNotFoundError:
        print("Error: the file doesn't exists")


main()
